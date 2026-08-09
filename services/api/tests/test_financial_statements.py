from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app.services.financial_statements import (
    build_statement_csv,
    build_statement_workbook,
    normalize_statement_frame,
)


def _payload(statement: dict) -> dict:
    return {
        "ticker": "TEST",
        "market_data_ticker": "TEST.NS",
        "currency": "INR",
        "available": True,
        "statements": [statement],
        "source": "Yahoo Finance",
        "source_url": "https://finance.yahoo.com/quote/TEST.NS/financials",
        "retrieved_at": "2026-07-29T10:00:00Z",
        "is_delayed_or_unverified": True,
        "limitations": ["Delayed standardized data.", "Missing values remain blank."],
    }


def test_statement_normalization_uses_analyst_order_and_typed_rows():
    frame = pd.DataFrame(
        {
            pd.Timestamp("2025-03-31"): [0.25, 20_000_000, 100_000_000],
            pd.Timestamp("2024-03-31"): [0.20, float("nan"), 80_000_000],
        },
        index=["Tax Rate For Calcs", "Net Income", "Total Revenue"],
    )

    statement = normalize_statement_frame(frame, "income", "annual")

    assert statement["periods"] == ["2024-03-31", "2025-03-31"]
    assert [row["label"] for row in statement["rows"]] == [
        "Total Revenue",
        "Net Income",
        "Tax Rate For Calcs",
    ]
    assert statement["rows"][0]["is_total"] is True
    assert statement["rows"][-1]["value_type"] == "percent"
    assert statement["rows"][1]["values"]["2024-03-31"] is None


def test_csv_export_preserves_numbers_and_labels_units():
    statement = {
        "statement_type": "income",
        "title": "Income Statement",
        "frequency": "annual",
        "periods": ["2024-03-31", "2025-03-31"],
        "rows": [
            {
                "key": "Total Revenue",
                "label": "Total Revenue",
                "value_type": "currency",
                "is_total": True,
                "values": {"2024-03-31": 80_000_000, "2025-03-31": 100_000_000},
            },
            {
                "key": "Tax Rate For Calcs",
                "label": "Tax Rate For Calcs",
                "value_type": "percent",
                "is_total": False,
                "values": {"2024-03-31": 0.2, "2025-03-31": 0.25},
            },
        ],
    }

    csv_text = build_statement_csv(_payload(statement), "income", "annual").decode(
        "utf-8-sig"
    )

    assert "Line item,Unit,2024-03-31,2025-03-31" in csv_text
    assert "Total Revenue,INR crore,8.0,10.0" in csv_text
    assert "Tax Rate For Calcs,%,20.0,25.0" in csv_text


def test_xlsx_export_has_cover_formatting_and_numeric_statement_values():
    statement = {
        "statement_type": "income",
        "title": "Income Statement",
        "frequency": "annual",
        "periods": ["2024-03-31", "2025-03-31"],
        "rows": [
            {
                "key": "Total Revenue",
                "label": "Total Revenue",
                "value_type": "currency",
                "is_total": True,
                "values": {"2024-03-31": 80_000_000, "2025-03-31": 100_000_000},
            },
            {
                "key": "Tax Rate For Calcs",
                "label": "Tax Rate For Calcs",
                "value_type": "percent",
                "is_total": False,
                "values": {"2024-03-31": 0.2, "2025-03-31": 0.25},
            },
        ],
    }

    workbook_bytes = build_statement_workbook(_payload(statement))
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)

    assert workbook.sheetnames == ["Cover", "Annual Income"]
    assert workbook["Cover"]["A1"].value == "TEST — Financial Statements"
    sheet = workbook["Annual Income"]
    assert sheet.freeze_panes == "C5"
    assert sheet["C5"].value == 8
    assert sheet["D6"].value == 0.25
    assert sheet["D6"].number_format.startswith("0.0%")
    assert sheet.sheet_view.showGridLines is False


def test_xlsx_export_explains_empty_statement_frequency():
    statement = {
        "statement_type": "cash_flow",
        "title": "Cash Flow Statement",
        "frequency": "quarterly",
        "periods": [],
        "rows": [],
    }

    workbook_bytes = build_statement_workbook(_payload(statement))
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert "No populated source values" in workbook["Quarterly Cash Flow"]["A5"].value
