"""Full standardized financial statements and analyst-friendly exports.

The development data source is yfinance. Values are kept numeric and retain
their source currency. This module does not estimate, interpolate, or ask an
LLM to fill missing statement values.
"""

import asyncio
import csv
import io
import math
from datetime import UTC, datetime
from typing import Any, Literal
from urllib.parse import quote

from app.core.cache import FileCache, cache_key
from app.core.config import get_settings

StatementType = Literal["income", "balance_sheet", "cash_flow"]
StatementFrequency = Literal["annual", "quarterly"]

STATEMENT_TITLES: dict[StatementType, str] = {
    "income": "Income Statement",
    "balance_sheet": "Balance Sheet",
    "cash_flow": "Cash Flow Statement",
}
STATEMENT_TTL_SECONDS = 6 * 60 * 60

PERCENT_ROWS = {
    "Tax Rate For Calcs",
}
PER_SHARE_MARKERS = (
    " EPS",
    "Per Share",
)
SHARE_COUNT_MARKERS = (
    "Average Shares",
    "Shares Number",
    "Share Issued",
)
TOTAL_ROW_LABELS = {
    "Total Revenue",
    "Gross Profit",
    "Operating Income",
    "EBIT",
    "EBITDA",
    "Pretax Income",
    "Net Income",
    "Total Assets",
    "Total Current Assets",
    "Current Assets",
    "Total Non Current Assets",
    "Total Liabilities Net Minority Interest",
    "Current Liabilities",
    "Stockholders Equity",
    "Total Equity Gross Minority Interest",
    "Operating Cash Flow",
    "Investing Cash Flow",
    "Financing Cash Flow",
    "Free Cash Flow",
    "End Cash Position",
}


def _finite_number(value: Any) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _row_value_type(label: str) -> Literal["currency", "percent", "per_share", "shares"]:
    if label in PERCENT_ROWS or "Tax Rate" in label:
        return "percent"
    if any(marker in label for marker in PER_SHARE_MARKERS):
        return "per_share"
    if any(marker in label for marker in SHARE_COUNT_MARKERS):
        return "shares"
    return "currency"


def _period_iso(value: Any) -> str:
    date_value = getattr(value, "date", None)
    if callable(date_value):
        return date_value().isoformat()
    return str(value)[:10]


def normalize_statement_frame(
    frame: Any,
    statement_type: StatementType,
    frequency: StatementFrequency,
) -> dict[str, Any]:
    """Convert a yfinance statement DataFrame into JSON-safe row-oriented data."""

    if frame is None or getattr(frame, "empty", True):
        return {
            "statement_type": statement_type,
            "title": STATEMENT_TITLES[statement_type],
            "frequency": frequency,
            "periods": [],
            "rows": [],
        }

    ordered_columns = sorted(frame.columns, key=_period_iso)
    periods = [_period_iso(column) for column in ordered_columns]
    rows: list[dict[str, Any]] = []
    # Yahoo returns these frames from derived metrics back toward the top line.
    # Reversing produces the conventional analyst reading order.
    for raw_label in reversed(list(frame.index)):
        label = str(raw_label).strip()
        values = {
            _period_iso(column): _finite_number(frame.at[raw_label, column])
            for column in ordered_columns
        }
        if not any(value is not None for value in values.values()):
            continue
        rows.append(
            {
                "key": label,
                "label": label,
                "value_type": _row_value_type(label),
                "is_total": label in TOTAL_ROW_LABELS or label.startswith("Total "),
                "values": values,
            }
        )

    populated_periods = [
        period
        for period in periods
        if any(row["values"].get(period) is not None for row in rows)
    ]
    for row in rows:
        row["values"] = {period: row["values"].get(period) for period in populated_periods}
    return {
        "statement_type": statement_type,
        "title": STATEMENT_TITLES[statement_type],
        "frequency": frequency,
        "periods": populated_periods,
        "rows": rows,
    }


def _fetch_financial_statements_sync(ticker: str, market_data_ticker: str) -> dict[str, Any]:
    import yfinance as yf

    stock = yf.Ticker(market_data_ticker)
    try:
        currency = getattr(stock.fast_info, "currency", None) or "INR"
    except Exception:
        currency = "INR"

    frames: tuple[tuple[StatementType, StatementFrequency, Any], ...] = (
        ("income", "annual", stock.income_stmt),
        ("income", "quarterly", stock.quarterly_income_stmt),
        ("balance_sheet", "annual", stock.balance_sheet),
        ("balance_sheet", "quarterly", stock.quarterly_balance_sheet),
        ("cash_flow", "annual", stock.cash_flow),
        ("cash_flow", "quarterly", stock.quarterly_cash_flow),
    )
    statements = [
        normalize_statement_frame(frame, statement_type, frequency)
        for statement_type, frequency, frame in frames
    ]
    available = any(statement["rows"] for statement in statements)
    retrieved_at = datetime.now(UTC).isoformat()
    return {
        "ticker": ticker,
        "market_data_ticker": market_data_ticker,
        "currency": currency,
        "available": available,
        "statements": statements,
        "source": "Yahoo Finance",
        "source_url": f"https://finance.yahoo.com/quote/{quote(market_data_ticker)}/financials",
        "retrieved_at": retrieved_at,
        "is_delayed_or_unverified": True,
        "limitations": [
            "These are delayed, standardized market-data statements and may differ from the company's statutory NSE filing presentation.",
            "Unavailable source values remain blank; no values are estimated or generated by AI.",
        ],
    }


async def get_full_financial_statements(
    ticker: str,
    market_data_ticker: str,
) -> dict[str, Any]:
    cache = FileCache(get_settings().cache_path, "financial_statements")
    key = cache_key("full-statements-v1", market_data_ticker.upper())
    cached = cache.get(key, STATEMENT_TTL_SECONDS)
    if cached is not None:
        return cached
    data = await asyncio.to_thread(
        _fetch_financial_statements_sync,
        ticker.upper(),
        market_data_ticker.upper(),
    )
    cache.put(key, data, source="yfinance")
    return data


def _display_unit(currency: str, value_type: str) -> str:
    if value_type == "percent":
        return "%"
    if value_type == "per_share":
        return f"{currency} / share"
    if value_type == "shares":
        return "shares"
    return f"{currency} crore" if currency == "INR" else f"{currency} millions"


def _export_value(value: float | None, value_type: str, currency: str) -> float | None:
    if value is None:
        return None
    if value_type == "percent":
        return value * 100
    if value_type in {"per_share", "shares"}:
        return value
    return value / (10_000_000 if currency == "INR" else 1_000_000)


def _statement_from_payload(
    payload: dict[str, Any],
    statement_type: StatementType,
    frequency: StatementFrequency,
) -> dict[str, Any]:
    return next(
        (
            statement
            for statement in payload["statements"]
            if statement["statement_type"] == statement_type
            and statement["frequency"] == frequency
        ),
        {
            "statement_type": statement_type,
            "title": STATEMENT_TITLES[statement_type],
            "frequency": frequency,
            "periods": [],
            "rows": [],
        },
    )


def build_statement_csv(
    payload: dict[str, Any],
    statement_type: StatementType,
    frequency: StatementFrequency,
) -> bytes:
    statement = _statement_from_payload(payload, statement_type, frequency)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        [
            "Line item",
            "Unit",
            *(period for period in statement["periods"]),
        ]
    )
    for row in statement["rows"]:
        writer.writerow(
            [
                row["label"],
                _display_unit(payload["currency"], row["value_type"]),
                *(
                    _export_value(
                        row["values"].get(period),
                        row["value_type"],
                        payload["currency"],
                    )
                    for period in statement["periods"]
                ),
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def _safe_sheet_title(statement: dict[str, Any]) -> str:
    prefix = "Annual" if statement["frequency"] == "annual" else "Quarterly"
    suffix = {
        "income": "Income",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow",
    }[statement["statement_type"]]
    return f"{prefix} {suffix}"


def build_statement_workbook(payload: dict[str, Any]) -> bytes:
    """Create a clean, source-labelled workbook containing all six statements."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    workbook.properties.title = f"{payload['ticker']} Financial Statements"
    workbook.properties.subject = "Historical financial statement export"
    workbook.properties.creator = "EquityLens"

    navy = "17233B"
    emerald = "047857"
    light_emerald = "E8F5EE"
    light_gray = "EEF1F4"
    border_gray = "D8DEE6"
    white = "FFFFFF"
    dark = "17202A"
    thin_gray = Side(style="thin", color=border_gray)

    cover.sheet_view.showGridLines = False
    cover.merge_cells("A1:F2")
    cover["A1"] = f"{payload['ticker']} — Financial Statements"
    cover["A1"].font = Font(name="Arial", size=18, bold=True, color=white)
    cover["A1"].fill = PatternFill("solid", fgColor=navy)
    cover["A1"].alignment = Alignment(vertical="center", horizontal="left")
    metadata = [
        ("Market-data ticker", payload["market_data_ticker"]),
        ("Currency", payload["currency"]),
        (
            "Currency scale",
            "INR crore" if payload["currency"] == "INR" else f"{payload['currency']} millions",
        ),
        ("Source", payload["source"]),
        ("Source URL", payload["source_url"]),
        ("Retrieved", payload["retrieved_at"]),
        ("Status", "Delayed / unverified market data"),
    ]
    for row_index, (label, value) in enumerate(metadata, start=4):
        cover.cell(row=row_index, column=1, value=label).font = Font(
            name="Arial",
            bold=True,
            color=dark,
        )
        cover.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=6,
        )
        cover.cell(row=row_index, column=2, value=value).font = Font(
            name="Arial",
            color=dark,
        )
    cover["A13"] = "Workbook contents"
    cover["A13"].font = Font(name="Arial", bold=True, color=white)
    cover["A13"].fill = PatternFill("solid", fgColor=emerald)
    cover.merge_cells("A13:F13")
    for row_index, statement in enumerate(payload["statements"], start=14):
        cover.cell(row=row_index, column=1, value=_safe_sheet_title(statement))
        cover.cell(
            row=row_index,
            column=2,
            value=f"{len(statement['rows'])} populated line items",
        )
    note_row = 22
    cover.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=6)
    cover.cell(
        row=note_row,
        column=1,
        value=" ".join(payload.get("limitations", [])),
    )
    cover.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    cover.cell(row=note_row, column=1).fill = PatternFill("solid", fgColor=light_gray)
    cover.column_dimensions["A"].width = 24
    for column in range(2, 7):
        cover.column_dimensions[get_column_letter(column)].width = 18
    cover.freeze_panes = "A4"

    for statement in payload["statements"]:
        sheet = workbook.create_sheet(_safe_sheet_title(statement))
        sheet.sheet_view.showGridLines = False
        periods = statement["periods"]
        last_column = max(2, 2 + len(periods))
        last_letter = get_column_letter(last_column)
        sheet.merge_cells(f"A1:{last_letter}2")
        frequency_label = statement["frequency"].title()
        sheet["A1"] = (
            f"{payload['ticker']} — {frequency_label} {statement['title']}"
        )
        sheet["A1"].font = Font(name="Arial", size=16, bold=True, color=white)
        sheet["A1"].fill = PatternFill("solid", fgColor=navy)
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.merge_cells(f"A3:{last_letter}3")
        sheet["A3"] = (
            f"Source: {payload['source']} | Retrieved: {payload['retrieved_at']} | "
            "Historical actuals; blanks indicate unavailable source values."
        )
        sheet["A3"].font = Font(name="Arial", size=9, italic=True, color="5D6673")
        sheet["A3"].alignment = Alignment(wrap_text=True)

        headers = ["Line item", "Unit", *periods]
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=column_index, value=header)
            cell.font = Font(name="Arial", bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=emerald)
            cell.alignment = Alignment(
                horizontal="right" if column_index > 2 else "left",
                vertical="center",
            )
        sheet.row_dimensions[4].height = 24

        for row_index, row in enumerate(statement["rows"], start=5):
            label_cell = sheet.cell(row=row_index, column=1, value=row["label"])
            unit_cell = sheet.cell(
                row=row_index,
                column=2,
                value=_display_unit(payload["currency"], row["value_type"]),
            )
            label_cell.font = Font(name="Arial", color=dark, bold=row["is_total"])
            unit_cell.font = Font(name="Arial", size=9, color="697386")
            for period_index, period in enumerate(periods, start=3):
                raw_value = row["values"].get(period)
                exported = _export_value(
                    raw_value,
                    row["value_type"],
                    payload["currency"],
                )
                cell = sheet.cell(row=row_index, column=period_index, value=exported)
                cell.font = Font(name="Arial", color=dark, bold=row["is_total"])
                cell.alignment = Alignment(horizontal="right")
                if row["value_type"] == "percent":
                    # The exported CSV uses percentage points, but Excel keeps
                    # rates as decimals so standard percentage formats remain live.
                    cell.value = raw_value
                    cell.number_format = "0.0%;[Red](0.0%);-"
                elif row["value_type"] == "per_share":
                    cell.number_format = "0.00;[Red](0.00);-"
                elif row["value_type"] == "shares":
                    cell.number_format = "#,##0;[Red](#,##0);-"
                else:
                    cell.number_format = "#,##0.00;[Red](#,##0.00);-"
                if period_index == last_column:
                    cell.fill = PatternFill("solid", fgColor=light_emerald)

            if row["is_total"]:
                for column_index in range(1, last_column + 1):
                    sheet.cell(row=row_index, column=column_index).border = Border(
                        top=thin_gray
                    )

        if not statement["rows"]:
            sheet.merge_cells(
                start_row=5,
                start_column=1,
                end_row=6,
                end_column=last_column,
            )
            sheet["A5"] = "No populated source values are available for this statement frequency."
            sheet["A5"].font = Font(name="Arial", italic=True, color="697386")
            sheet["A5"].fill = PatternFill("solid", fgColor=light_gray)
            sheet["A5"].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        sheet.column_dimensions["A"].width = 52
        sheet.column_dimensions["B"].width = 17
        for column_index in range(3, last_column + 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = 16
        sheet.freeze_panes = "C5"
        sheet.auto_filter.ref = f"A4:{last_letter}{4 + len(statement['rows'])}"

    binary = io.BytesIO()
    workbook.save(binary)
    return binary.getvalue()
